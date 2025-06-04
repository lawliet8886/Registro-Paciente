import sys
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter
from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QLineEdit, QPushButton, QMessageBox, QVBoxLayout, QWidget, QHBoxLayout, QCheckBox, QDialog, QFormLayout, QTimeEdit, QComboBox
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPixmap
from datetime import datetime
from PyQt5.QtCore import QTime

class EncaminhamentoDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Tipo de Encaminhamento")
        layout = QFormLayout()
        self.combobox_encaminhamento = QComboBox()
        encaminhamento_types = ["Demanda Espontânea", "Abordagem na Rua", "Abrigo", "Ambulatório", "Atenção Básica", "Caps da RAPS Municipal", "Caps de outro Município", "Conselho Tutelar", "Consultório na Rua", "CREAS/CRAS", "Escola", "Hospital Geral", "Hospital Psiquiátrico", "Justiça", "Rede Intersetorial", "Rede Privada Amb/Hospital", "Urgência/Emergência"]
        self.combobox_encaminhamento.addItems(encaminhamento_types)
        layout.addRow("Tipo de Encaminhamento:", self.combobox_encaminhamento)
        self.button_save = QPushButton("Salvar")
        layout.addRow("", self.button_save)
        self.setLayout(layout)
        self.button_save.clicked.connect(self.accept)

    def get_encaminhamento(self):
        return self.combobox_encaminhamento.currentText()
        
class TimeInputDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Intervalo de Convivência")
        layout = QFormLayout()
        self.timeedit_start = QTimeEdit()
        self.timeedit_start.setDisplayFormat("HH:mm")
        self.timeedit_end = QTimeEdit()
        self.timeedit_end.setDisplayFormat("HH:mm")
        layout.addRow("Horário Inicial:", self.timeedit_start)
        layout.addRow("Horário Final:", self.timeedit_end)
        self.button_save = QPushButton("Salvar")
        layout.addRow("", self.button_save)
        self.setLayout(layout)
        self.button_save.clicked.connect(self.accept)

    def get_time_interval(self):
        start_time = self.timeedit_start.time().toString("HH:mm")
        end_time = self.timeedit_end.time().toString("HH:mm")
        return start_time, end_time

class MyLineEdit(QLineEdit):
    def keyPressEvent(self, event):
        super().keyPressEvent(event)
        if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:
            self.focusNextChild()

class PatientRegistration(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Sistema de Registro de Pacientes")
        self.setGeometry(100, 100, 400, 300)
        self.start_time = None
        self.end_time = None
        self.encaminhamento = None
        self.ws_consolidados = None
        widget = QWidget(self)
        self.setCentralWidget(widget)
        layout = QVBoxLayout()
        widget.setLayout(layout)
        background = QLabel(self)
        pixmap = QPixmap("OIG.jpeg")
        scaled_pixmap = pixmap.scaled(50, 50, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        background.setPixmap(pixmap)
        background.setAlignment(Qt.AlignCenter)
        layout.addWidget(background)
        form_layout = QVBoxLayout()
        self.label_name = QLabel("Nome do Paciente:", self)
        self.textedit_name = MyLineEdit(self)
        self.textedit_name.setMaximumHeight(80)
        self.label_demand = QLabel("Tipo de Demanda:", self)
        form_layout.addWidget(self.label_demand)  # You add the label to the form_layout here
        demand_types = ["A", "R", "M", "AN", "C", "RM", "Grupos/Eventos", "Outros", "AI", "REA"]
        checkbox_rows = [demand_types[i:i+5] for i in range(0, len(demand_types), 5)]
        self.checkboxes = []
        checkbox_layouts = []
        for row in checkbox_rows:
            checkbox_layout = QHBoxLayout()
            checkbox_layouts.append(checkbox_layout)
            for demand_type in row:
                  checkbox = QCheckBox(demand_type, self)
                  self.checkboxes.append(checkbox)
                  checkbox_layout.addWidget(checkbox)
                  if demand_type in ["AI", "REA"]:
                     checkbox.stateChanged.connect(self.handle_ai_rea_checkboxes)
                  if demand_type == "C":
                     checkbox.stateChanged.connect(self.handle_c_checkbox) 
                  if demand_type == "AN":
                     checkbox.stateChanged.connect(self.handle_an_checkbox)
                                          
        for checkbox_layout in checkbox_layouts:
            form_layout.addLayout(checkbox_layout)            
        self.label_c_times = QLabel("", self)
        self.label_encaminhamento = QLabel("", self)
        self.label_reference = QLabel("Prof. de referência:", self)
        self.textedit_reference = MyLineEdit(self)
        self.textedit_reference.setMaximumHeight(80)        
        # Adicionando os novos checkboxes
        self.checkbox_lunch = QCheckBox("Almoço", self)
        self.checkbox_snack = QCheckBox("Lanche", self)
        self.checkbox_dinner = QCheckBox("Janta", self)
        # Adicionando o rótulo
        self.label_demand_choice = QLabel("Selecione após escolher o Tipo de Demanda(opcional):", self)
        form_layout.addWidget(self.label_demand_choice)  # Adicione este rótulo antes dos checkboxes       
        form_layout.addWidget(self.checkbox_lunch)
        form_layout.addWidget(self.checkbox_snack)
        form_layout.addWidget(self.checkbox_dinner)        
        form_layout.addWidget(self.label_name)
        form_layout.addWidget(self.textedit_name)
        form_layout.addWidget(self.label_c_times)
        form_layout.addWidget(self.label_encaminhamento)
        form_layout.addWidget(self.label_reference)
        form_layout.addWidget(self.textedit_reference)
        layout.addLayout(form_layout)

        # add label and text edit for observations
        self.label_observations = QLabel("Observações:", self)
        self.textedit_observations = MyLineEdit(self)
        self.textedit_observations.setMaximumHeight(80)
        form_layout.addWidget(self.label_observations)
        form_layout.addWidget(self.textedit_observations)
        self.button_register = QPushButton("Registrar", self)
        self.button_register.setMaximumWidth(150)
        self.button_register.setDefault(True)  # make the Register button default
        layout.addWidget(self.button_register)
        self.button_register.clicked.connect(self.register_patient) 

        if os.path.exists("pacientes_recepção.xlsx"):
            self.update_consolidated_sheet()
            self.update_consolidated_totals_sheet() 
        self.set_tab_order()  # set the tab order after all widgets are initialized

    def set_tab_order(self):
        self.setTabOrder(self.textedit_name, self.textedit_reference)
        self.setTabOrder(self.textedit_reference, self.textedit_observations)
        self.setTabOrder(self.textedit_observations, self.button_register)                                 
                                                              
# update handle_c_checkbox method
    def handle_c_checkbox(self, state):
        if state == Qt.Checked:
            time_dialog = TimeInputDialog(self)
            if time_dialog.exec_() == QDialog.Accepted:
                self.start_time, self.end_time = time_dialog.get_time_interval()
                self.label_c_times.    setText(f"Convivência das {self.start_time} até as {self.end_time}")
        else:
            self.label_c_times.setText("")

    def handle_an_checkbox(self, state):
        if state == Qt.Checked:
            self.checkbox_lunch.setChecked(True)
            self.checkbox_snack.setChecked(True)
            self.checkbox_dinner.setChecked(True)
        else:
            self.checkbox_lunch.setChecked(False)
            self.checkbox_snack.setChecked(False)
            self.checkbox_dinner.setChecked(False)            
                                    
    def handle_ai_rea_checkboxes(self):
        checked = False
        for cb in self.checkboxes:
            if cb.text() in ["AI", "REA"] and cb.isChecked():
                encaminhamento_dialog = EncaminhamentoDialog(self)
                if encaminhamento_dialog.exec_() == QDialog.Accepted:
                    self.encaminhamento = encaminhamento_dialog.get_encaminhamento()
                    self.label_encaminhamento.setText(f"Tipo de Encaminhamento: {self.encaminhamento}")
                checked = True
        if not checked:
            self.encaminhamento = None
            self.label_encaminhamento.setText("") 

                                                            
    def register_patient(self):
        try:
            patient_name = self.textedit_name.text().strip()
            reference_prof = self.textedit_reference.text().strip()
            current_date = datetime.now().strftime("%d/%m/%Y")
            current_time = datetime.now().strftime("%H:%M")
            demands = [cb.text() for cb in self.checkboxes if cb.isChecked()]
            observations = self.textedit_observations.text().strip()  # Get the text from the Observations field
            
             # Check if Observations field is empty
            if not reference_prof:
                QMessageBox.warning(self, "Aviso", "Por favor, insira o profissional de referência ou de atendimento.")
                return           

            if not patient_name or not demands:
                QMessageBox.warning(self, "Aviso", "Por favor, insira o nome do paciente e selecione ao menos uma demanda.")
                return
            if "C" in demands and self.start_time and self.end_time:
                demands[demands.index("C")] = f"C ({self.start_time}-{self.end_time})"

            if not os.path.exists("pacientes_recepção.xlsx"):
                wb = Workbook()
                ws1 = wb.active
                ws1.title = "Pacientes"
                ws1.append(["Nome do Paciente", "Tipo de Demanda", "Prof. de referência", "Data", "Hora", "Observações"])

                ws2 = wb.create_sheet("Almoço")
                ws2.append(["Nome do Paciente", "Tipo de Demanda", "Prof. de referência", "Data", "Hora", "Observações"])

                ws3 = wb.create_sheet("Janta")
                ws3.append(["Nome do Paciente", "Tipo de Demanda", "Prof. de referência", "Data", "Hora", "Observações"])

                ws4 = wb.create_sheet("Acolhimentos")
                ws4.append(["Nome do Paciente", "Tipo de Demanda","Tipo de encaminhamento",  "Prof. de referência", "Data", "Hora", "Observações"])
                ws5 = wb.create_sheet("Lanche")
                ws5.append(["Nome do Paciente", "Tipo de Demanda", "Prof. de referência", "Data", "Hora", "Observações"])                
            else:
                wb = load_workbook("pacientes_recepção.xlsx")
                ws1 = wb["Pacientes"]
                ws2 = wb["Almoço"]                
                ws3 = wb["Janta"]
                ws4 = wb["Acolhimentos"]
                ws5 = wb["Lanche"]

            if "AI" in demands or "REA" in demands:
                ws4.append([patient_name, ", ".join(demands), self.encaminhamento,reference_prof, current_date, current_time, observations])
            else:
                ws1.append([patient_name, ", ".join(demands), reference_prof, current_date, current_time, observations])

            if "C" in demands and self.start_time is not None and self.end_time is not None:
                start_time = QTime.fromString(self.start_time, "HH:mm")
                end_time = QTime.fromString(self.end_time, "HH:mm")
                              
            if self.checkbox_lunch.isChecked():
                ws2.append([patient_name, ", ".join(demands), reference_prof, current_date, current_time, observations])                
            if self.checkbox_dinner.isChecked():
                ws3.append([patient_name, ", ".join(demands), reference_prof, current_date, current_time, observations])
            if self.checkbox_snack.isChecked():
                ws5.append([patient_name, ", ".join(demands), reference_prof, current_date, current_time, observations])                
            for ws in [ws1, ws2, ws3, ws4, ws5]:
                for column_cells in ws.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length
                    
            # Adicione filtros para o dia atual em todas as planilhas
            current_date_str = datetime.now().strftime("%d/%m/%Y")
            filter_column_index = {
                ws1: 3,
                ws2: 3,
                ws3: 3,
                ws4: 4,
                ws5: 3 
                               
            }
            for ws in [ws1, ws2, ws3, ws4, ws5]:
                ws.auto_filter.ref = ws.dimensions
                column_index = filter_column_index.get(ws, 3)
                ws.auto_filter.add_filter_column(column_index, [current_date_str])
                
            wb.save("pacientes_recepção.xlsx")
            self.update_consolidated_sheet()
            self.update_consolidated_totals_sheet() 
            self.clear_fields()
        except Exception as e:
            QMessageBox.critical(self, "Erro", str(e))                    
    def update_consolidated_sheet(self):
        wb = load_workbook("pacientes_recepção.xlsx")
        current_date = datetime.now().strftime("%d-%m-%Y")
        sheet_name = f"Consolidados"
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(["Data", "Pacientes para o Almoço","Pacientes para o Lanche", "Pacientes para a Janta", "Total de Pacientes", "Total de Acolhimentos", "Analise Consolidados"])
        else:
        	ws = wb[sheet_name]

        current_date_str = datetime.now().        strftime("%d/%m/%Y")  # formato da data deve corresponder ao da planilha
        lunch_count = sum(1 for row in   wb["Almoço"].iter_rows(min_row=2, values_only=True) if row[3] == current_date_str)
        snack_count = sum(1 for row in wb["Lanche"].iter_rows(min_row=2, values_only=True) if row[3] == current_date_str)
        dinner_count = sum(1 for row in wb["Janta"].iter_rows(min_row=2, values_only=True) if row[3] == current_date_str)
        total_patients = sum(1 for row in wb["Pacientes"].iter_rows(min_row=2, values_only=True) if row[3] == current_date_str)
        total_acolhimentos = sum(1 for row in wb["Acolhimentos"].iter_rows(min_row=2, values_only=True) if row[4] == current_date_str)
    # Check if current date already exists in the sheet
        for row in ws.iter_rows(min_row=2, 		max_row=ws.max_row, max_col=1):
            if current_date == row[0].value:
            # If current date exists, update the row
            	ws.cell(row=row[0].row, column=2, value=lunch_count)
            	ws.cell(row=row[0].row, column=3, value=snack_count)            	            	
            	ws.cell(row=row[0].row, column=4, value=dinner_count)
            	ws.cell(row=row[0].row, column=5, value=total_patients)
            	ws.cell(row=row[0].row, column=6, value=total_acolhimentos)           	
            	break
        else:
        # If current date does not exist, append a new row
        	ws.append([current_date, lunch_count, snack_count, dinner_count, total_patients, 	total_acolhimentos])        	
         # Adicione a análise dos tipos de demanda à coluna "Análise Consolidados"
        demand_analysis = self.get_demand_analysis(wb)
        ws.cell(row=ws.max_row, column=7, value=demand_analysis)       	

        for column_cells in ws.columns:
        	length = max(len(str(cell.value)) for cell in column_cells)
        	ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length

# Aplicar filtro de data em todas as planilhas
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            date_column = 3 if sheet in ["Almoço", "Lanche", "Janta", "Pacientes"] else 4 if sheet == "Acolhimentos" else 0
            ws.auto_filter.ref = ws.dimensions
            ws.auto_filter. add_filter_column(date_column, [current_date])

        wb.save("pacientes_recepção.xlsx")
        QMessageBox.information(self, "Sucesso", "Atualizado com sucesso!")
       
    def update_consolidated_totals_sheet(self):
        wb = load_workbook("pacientes_recepção.xlsx")
        current_date = datetime.now().strftime("%d-%m-%Y")
        sheet_name = f"Consolidados Totais"
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(["Data", "Pacientes para o Almoço","Pacientes para o Lanche", "Pacientes para a Janta", "Total de Pacientes", "Total de Acolhimentos", "Analise Consolidados"])
        else:
            ws = wb[sheet_name]

        lunch_count = sum(1 for row in wb["Almoço"].iter_rows(min_row=2, values_only=True))
        snack_count = sum(1 for row in wb["Lanche"].iter_rows(min_row=2, values_only=True))
        dinner_count = sum(1 for row in wb["Janta"].iter_rows(min_row=2, values_only=True))
        total_patients = sum(1 for row in wb["Pacientes"].iter_rows(min_row=2, values_only=True))
        total_acolhimentos = sum(1 for row in wb["Acolhimentos"].iter_rows(min_row=2, values_only=True))

        if ws.max_row > 1: # if there are more than one rows in the sheet (i.e., data have been inserted before)
        # update the counts in the first row
            ws.cell(row=2, column=2, value=lunch_count)
            ws.cell(row=2, column=3, value=snack_count)
            ws.cell(row=2, column=4, value=dinner_count)
            ws.cell(row=2, column=5, value=total_patients)
            ws.cell(row=2, column=6, value=total_acolhimentos)
        else: # if there are not more than one rows (i.e., it's the first time we're adding data)
            ws.append(["Totais", lunch_count, snack_count, dinner_count, total_patients, total_acolhimentos])

        demand_analysis = self.get_demand_analysis(wb, total_analysis=True)
        ws.cell(row=ws.max_row, column=7, value=demand_analysis)

    # rest of your function...
       

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            date_column = 3 if sheet in ["Almoço", "Lanche", "Janta", "Pacientes"] else 4 if sheet == "Acolhimentos" else 0
            ws.auto_filter.ref = ws.dimensions
            ws.auto_filter.add_filter_column(date_column, [current_date])

        wb.save("pacientes_recepção.xlsx")
              
    def get_demand_analysis(self, wb, total_analysis=False):
        current_date_str = datetime.now().strftime("%d/%m/%Y")  # formato da data deve corresponder ao da planilha
        demand_analysis = ""
        for sheet_name in ["Pacientes", "Almoço", "Lanche", "Janta", "Acolhimentos"]:
            ws = wb[sheet_name]
            demand_counts = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                date_index = 4 if sheet_name == "Acolhimentos" else 3
                if not total_analysis and row[date_index] != current_date_str:
                    continue  # skip rows that are not from the current date
                demands = row[1].split(", ")
                for demand in demands:
                    if demand in demand_counts:
                        demand_counts[demand] += 1
                    else:
                        demand_counts[demand] = 1
            demand_analysis += f"{sheet_name}:\n"
            for demand, count in demand_counts.items():
                demand_analysis += f"{demand}: {count}\n"
            demand_analysis += "\n"

        return demand_analysis
                
    def clear_fields(self):
        self.textedit_name.clear()
        self.textedit_reference.clear()
        self.textedit_observations.clear()        
        for cb in self.checkboxes:
            cb.setChecked(False)
        self.start_time = None
        self.end_time = None
        self.label_c_times.setText("")
        self.label_encaminhamento.setText("")
        self.checkbox_lunch.setChecked(False)
        self.checkbox_dinner.setChecked(False)
        self.checkbox_snack.setChecked(False)
        for cb in self.checkboxes:
            if cb.text() == "AN":
                cb.setChecked(False)
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = PatientRegistration()
    window.show()
    sys.exit(app.exec_())

